#from django.shortcuts import render
from django.http import HttpResponse
from .models import Contact, Campaign, CampaignLead, FunnelEvent

from datetime import datetime, time, timedelta
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
import json


def save_virtual_workbook(workbook):
    """Return an in-memory workbook, suitable for a Django response."""
    temp_buffer = BytesIO()
    try:
        archive = ZipFile(temp_buffer, 'w', ZIP_DEFLATED)
        writer = ExcelWriter(workbook, archive)
        writer.write_data() #archive)
    finally:
        archive.close()
    virtual_workbook = temp_buffer.getvalue()
    temp_buffer.close()
    return virtual_workbook

def eventlog(request, cid):
    try:
        cid=int(cid)
        campaign = Campaign.objects.get(pk=int(cid))
        xl = Workbook(write_only = True)
        ws = xl.create_sheet()

        ws.append(['Timestamp', 'Contact', 'Contact ID', 'Lead ID', 'Event Source', 'Action', 'Extra Data'])
        for ev in FunnelEvent.objects.prefetch_related('lead__contact').filter(seen=True, lead__campaign_id=cid):
            ws.append([ev.timestamp.replace(tzinfo=None), ev.lead.contact.msisdn, ev.lead.contact_id, ev.lead.pk, ev.source, ev.action, json.dumps(ev.data)])

        response = HttpResponse(content=save_virtual_workbook(xl), content_type=xl.mime_type)
        response['Content-Disposition'] = f"attachment; filename=events_cid{cid}.xlsx"
        return response
    except Exception as e:
        return HttpResponse(e.args)

def leadstatus(request, cid):
    try:
        cid=int(cid)
        campaign = Campaign.objects.get(pk=int(cid))
        xl = Workbook(write_only = True)
        ws = xl.create_sheet()

        ws.append(['Contact', 'Contact ID', 'Lead ID', 'Status', 'Stage', 'Tags', 'State'])
        for lead in CampaignLead.objects.prefetch_related('contact').filter(campaign_id=cid):
            ws.append([lead.contact.msisdn, lead.contact_id, lead.pk, CampaignLead.Status(lead.status).name, lead.stage, 
                ','.join(lead.contact.tags.all().values_list('name', flat=True).order_by('name')),
                json.dumps(lead.state, ensure_ascii=False)])

        response = HttpResponse(content=save_virtual_workbook(xl), content_type=xl.mime_type)
        response['Content-Disposition'] = f"attachment; filename=leads_cid{cid}.xlsx"
        return response
    except Exception as e:
        return HttpResponse(e.args)

def do_msgstat(msgs, lead):
    total = 0
    stat = { m:0 for m in msgs }
    for ev in lead.events.filter(source='teaser.tasks', action='status'):
        if ev.data['status'] != 'DELIVERED' or not ev.parent:
            continue
        total = total + 1
        if not 'text_id' in ev.parent.data:
            continue
        m = ev.parent.data['text_id']
        stat[m] = stat.get(m, 0) + 1
    return [ stat[m] for m in msgs ] + [ total ]

def do_lastopened(lead, tz = False):
    d = None
    try:
        ev = lead.events.filter(source='usher.shortener', action='follow').latest()
        d = ev.timestamp
        if not tz:
            d = d.replace(tzinfo=None)
    except Exception as e:
        pass
    return d

def count_opened_in(lead, start, days):
    if not start:
        return 0
    end = datetime.combine(start.date(), time.min, start.tzinfo)
    end = end + timedelta(days=days+1)
    return lead.events.filter(source='usher.shortener', action='follow', timestamp__gt=start, timestamp__lt=end).count()

def do_lastpaid(lead, tz = False):
    d = None
    try:
        ev = lead.events.filter(source='hooker.views', action='postback').latest()
        d = ev.timestamp
        if not tz:
            d = d.replace(tzinfo=None)
    except Exception as e:
        pass
    return d

def do_lastsendout(lead, tz = False):
    d = None
    try:
        ev = lead.events.filter(action='init').latest()
        ev = lead.events.filter(action='send_sms', timestamp__gt=ev.timestamp).earliest()
        d = ev.timestamp
        if not tz:
            d = d.replace(tzinfo=None)
    except Exception as e:
        pass
    return d

def leadstatus2(request, cid):
    try:
        cid=int(cid)
        campaign = Campaign.objects.get(pk=int(cid))

        msgs = sorted(list(campaign.settings['texts'].keys()))

        xl = Workbook(write_only = True)
        ws = xl.create_sheet()

        ws.append(['Contact', 'Contact ID', 'Lead ID', 'Status', 'Stage', 'Tags'] + msgs + ['Total', 'Last OPENED'])
        for lead in CampaignLead.objects.prefetch_related('contact').filter(campaign_id=cid):
            ws.append([lead.contact.msisdn, lead.contact_id, lead.pk, CampaignLead.Status(lead.status).name, lead.stage, 
                ','.join(lead.contact.tags.all().values_list('name', flat=True).order_by('name'))]
                + do_msgstat(msgs, lead) + [ do_lastopened(lead) ])

        response = HttpResponse(content=save_virtual_workbook(xl), content_type=xl.mime_type)
        response['Content-Disposition'] = f"attachment; filename=leads2_cid{cid}.xlsx"
        return response
    except Exception as e:
        return HttpResponse(e.args)

def leadstatus3(request, cid):
    try:
        cid=int(cid)
        campaign = Campaign.objects.get(pk=int(cid))

        xl = Workbook(write_only = True)
        ws = xl.create_sheet()

        ws.append(['Contact', 'LastOpenDate', 'Strategy', 'LastPayDate', 'Score', 'LastSendoutDate', 'OpenIn3', 'OpenIn6', 'OpenIn12'])

        for lead in CampaignLead.objects.prefetch_related('contact').filter(campaign_id=cid):
            last_open = do_lastopened(lead, True)
            last_pay = do_lastpaid(lead, True)
            last_sendout = do_lastsendout(lead, True)

            open_in3 = count_opened_in(lead, last_sendout, 3)
            open_in6 = count_opened_in(lead, last_sendout, 6)
            open_in12 = count_opened_in(lead, last_sendout, 12)

            strategy = 'Unknown'
            score = 0
            if lead.status == CampaignLead.Status.DEAD:
                strategy = 'Hold'

            try:
                lastDeliverEvent = lead.events.filter(action="status", data__status="Delivered", timestamp__gt=last_sendout).latest()
                last_deliver = lastDeliverEvent.timestamp
            except (FunnelEvent.DoesNotExist, ValueError):
                last_deliver = None

            if last_deliver and ((not last_open) or (last_open < last_deliver)):
                strategy = 'Uzer(0)'
                score = score + 1

            if last_sendout and last_open:
                sof = datetime.combine(last_sendout.date(), time.min, last_sendout.tzinfo)
                sof = sof + timedelta(days=3)
                tot = 0
                ntot = True
                while sof < last_open:
                    num = lead.events.filter(action="follow", timestamp__gt=last_sendout, timestamp__lt=sof+timedelta(days=1)).count()
                    sof = sof + timedelta(days=1)
                    if num <= 1:
                        tot  = tot + num
                    else:
                       ntot = False
                if tot > 0 and ntot:
                    uzerMetric = tot
                    strategy = 'Uzer({0})'.format(tot)
                    score = score + 2

            if last_sendout and (not last_pay or last_pay < last_sendout):
                num = lead.events.filter(action="follow", timestamp__gt=last_sendout).count()
                if num > 3:
                    strategy = 'aUzer({0})'.format(num)
                    score = score + 5

            if last_sendout and last_pay and (last_pay > last_sendout):
                strategy = 'Pay'
                score = score + 10

            ws.append([lead.contact.msisdn,
                last_open.replace(tzinfo=None).date() if last_open else None,
                strategy,
                last_pay.replace(tzinfo=None).date() if last_pay else None,
                score,
                last_sendout.replace(tzinfo=None).date() if last_sendout else None,
                open_in3, open_in6, open_in12])

        response = HttpResponse(content=save_virtual_workbook(xl), content_type=xl.mime_type)
        response['Content-Disposition'] = f"attachment; filename=leads3_cid{cid}.xlsx"
        return response

    except Exception as e:
        return HttpResponse(e.args)
